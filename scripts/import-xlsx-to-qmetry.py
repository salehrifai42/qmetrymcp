#!/usr/bin/env python3
"""
Import test cases from an Excel folder tree into QMetry.

For each `.xlsx` file under --input, sheet `E2E_TestCases` is parsed (column
positions auto-detected from the header row). Each row whose ID column looks
like an E2E ID starts a new test case; following rows are step continuations.

QMetry folder tree mirrors the input directory layout under --parent-folder-id:

    <parent>/
      <root-folder-name>/         (defaults to the basename of --input)
        <subdirectory>/
          <workbook-stem>/        (filename minus _E2E_TestCases)
            <test cases>

Idempotent: created test cases are appended to --log. Re-runs skip
(workbook, e2e_id) pairs already logged.

Env:
  QTM4J_API_KEY   required
  QTM4J_REGION    optional (US default, or AU)

Usage examples:
  # Full import (uses tenant defaults — CBS 2, Refactored_E2E, FS project)
  QTM4J_API_KEY=... python scripts/import-xlsx-to-qmetry.py

  # Different input folder, different parent
  QTM4J_API_KEY=... python scripts/import-xlsx-to-qmetry.py \\
      --input "Input/Cloud Network" --parent-folder-id 1234567

  # Dry-run to see what would be created
  python scripts/import-xlsx-to-qmetry.py --dry-run

  # Re-write precondition + steps for previously imported cases (use after
  # the parser changed or you discovered a column-mapping mistake)
  QTM4J_API_KEY=... python scripts/import-xlsx-to-qmetry.py --fix-existing
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl

# ── Tenant defaults (Generic Project project on US region) ────────────────
DEFAULT_PROJECT_ID = 10000
DEFAULT_PARENT_FOLDER_ID = 0       # Refactored_E2E
DEFAULT_INPUT_DIR = Path(__file__).parent.parent / "Input" / "CBS 2"
DEFAULT_COMPONENT_IDS = [0]         # CBS
DEFAULT_STATUS_ID = 0               # Done
DEFAULT_API_TEST_FIELD_ID = "qcf_0"
DEFAULT_API_TEST_OPTION_YES = "0"
DEFAULT_API_TEST_OPTION_NO = "0"

# Workbook P-code → QMetry priority id (project 10000)
PRIORITY_MAP = {
    "P1": 0,  # Blocker
    "P2": 0,  # High
    "P3": 0,  # Medium
    "P4": 0,  # Low
}
DEFAULT_PRIORITY_ID = 0  # Medium

# Header aliases — extend as new workbook layouts appear. Lower-case match.
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "e2e_id":    ("e2e test id", "e2e tc id", "test id", "tc id"),
    "summary":   ("e2e test name", "title", "summary", "test case name"),
    "category":  ("category",),
    "test_type": ("test type", "type"),
    "priority":  ("priority",),
    "notes":     ("notes / endpoint", "notes/endpoint", "notes", "endpoint", "precondition"),
    "step_num":  ("step #", "step no", "step number"),
    "action":    ("step summary (action)", "step summary", "step", "action", "step description"),
    "expected":  ("expected result", "expected", "expected outcome"),
}

E2E_ID_RE = re.compile(r"^([A-Z][A-Z0-9]*-){1,}\d+$", re.IGNORECASE)
TC_SUMMARY_MAX = 255

# ── HTTP ──────────────────────────────────────────────────────────────────────

class _Cfg:
    """Filled in main() from CLI args + env. Avoids module-globals churn."""
    api_key: str
    api_base: str

CFG = _Cfg()


def qtm(path: str, method: str = "GET", body=None, attempt: int = 1):
    req = urllib.request.Request(
        f"{CFG.api_base}{path}",
        method=method,
        headers={"apiKey": CFG.api_key, "Content-Type": "application/json"},
        data=json.dumps(body).encode() if body is not None else None,
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            raw = resp.read()
            return json.loads(raw) if raw else None
    except urllib.error.HTTPError as e:
        if e.code == 429 and attempt < 3:
            time.sleep(2 ** attempt)
            return qtm(path, method, body, attempt + 1)
        raise SystemExit(f"HTTP {e.code} on {method} {path}: {e.read().decode()[:400]}")


# ── Folder helpers ────────────────────────────────────────────────────────────

def _find_in_tree(node, target_id):
    if isinstance(node, list):
        for n in node:
            hit = _find_in_tree(n, target_id)
            if hit:
                return hit
        return None
    if not isinstance(node, dict):
        return None
    if node.get("id") == target_id:
        return node
    for child in node.get("children", []) or []:
        hit = _find_in_tree(child, target_id)
        if hit:
            return hit
    return None


def find_or_create_folder(project_id: int, name: str, parent_id: int) -> int:
    tree = qtm(f"/projects/{project_id}/testcase-folders")
    nodes = tree.get("data") if isinstance(tree, dict) and "data" in tree else tree
    target = _find_in_tree(nodes, parent_id)
    if target:
        for child in target.get("children", []) or []:
            if child.get("name") == name:
                return child["id"]
    created = qtm(
        f"/projects/{project_id}/testcase-folders",
        method="POST",
        body={"folderName": name, "parentId": parent_id},
    )
    return created["id"]


# ── Workbook parsing ──────────────────────────────────────────────────────────

def _is_valid_e2e_id(s: str) -> bool:
    s = s.strip()
    return bool(E2E_ID_RE.match(s)) and "E2E" in s.upper().split("-")


def _detect_columns(header_row) -> dict[str, int]:
    cols: dict[str, int] = {}
    for i, h in enumerate(header_row):
        h_low = str(h or "").strip().lower()
        for field, names in HEADER_ALIASES.items():
            if h_low in names and field not in cols:
                cols[field] = i
    return cols


def parse_workbook(xlsx_path: Path):
    """Yield test case dicts. Schema-aware via header aliases."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if "E2E_TestCases" not in wb.sheetnames:
        return
    ws = wb["E2E_TestCases"]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return
    cols = _detect_columns(header)
    if "e2e_id" not in cols or "action" not in cols:
        print(f"  ! unrecognized schema in {xlsx_path.name} — headers: {list(header)}", file=sys.stderr)
        return

    def cell(row, key):
        i = cols.get(key)
        return row[i] if i is not None and i < len(row) else None

    current = None
    for row in rows_iter:
        if not row or all(c is None for c in row):
            continue
        raw_id = cell(row, "e2e_id")
        e2e_id = raw_id.strip() if isinstance(raw_id, str) else (raw_id or "")
        if e2e_id and isinstance(e2e_id, str) and not _is_valid_e2e_id(e2e_id):
            if current:
                yield current
                current = None
            continue
        if e2e_id:
            if current:
                yield current
            pri_raw = cell(row, "priority")
            current = {
                "e2e_id": e2e_id,
                "summary": (cell(row, "summary") or "").strip()
                           if isinstance(cell(row, "summary"), str)
                           else str(cell(row, "summary") or ""),
                "priority_id": PRIORITY_MAP.get(str(pri_raw or "").strip().upper(), DEFAULT_PRIORITY_ID),
                "notes": cell(row, "notes") or "",
                "category": cell(row, "category") or "",
                "test_type": cell(row, "test_type") or "",
                "steps": [],
            }
        if not current:
            continue
        action = cell(row, "action")
        expected = cell(row, "expected")
        if action or expected:
            current["steps"].append({
                "stepDetails": str(action or "").strip(),
                "expectedResult": str(expected or "").strip(),
            })
    if current:
        yield current


# ── Log ───────────────────────────────────────────────────────────────────────

def load_log(log_path: Path) -> set[tuple[str, str]]:
    seen: set[tuple[str, str]] = set()
    if log_path.exists():
        with log_path.open() as f:
            for row in csv.DictReader(f):
                seen.add((row["workbook"], row["e2e_id"]))
    return seen


def append_log(log_path: Path, workbook, e2e_id, key, internal_id, status, error=""):
    new_file = not log_path.exists()
    with log_path.open("a", newline="") as f:
        w = csv.writer(f)
        if new_file:
            w.writerow(["workbook", "e2e_id", "qmetry_key", "internal_id", "status", "error"])
        w.writerow([workbook, e2e_id, key, internal_id, status, error])


# ── Modes ─────────────────────────────────────────────────────────────────────

def fix_existing(args, log_path: Path):
    """Re-write precondition + steps on previously imported cases."""
    if not log_path.exists():
        sys.exit(f"No log to fix: {log_path}")
    rows = [r for r in csv.DictReader(open(log_path)) if r["status"] == "ok"]
    by_wb: dict[str, list] = {}
    for r in rows:
        by_wb.setdefault(r["workbook"], []).append(r)

    fixed = failed = 0
    for wb_name, entries in by_wb.items():
        matches = list(args.input.rglob(wb_name))
        if not matches:
            print(f"  ! source missing: {wb_name}")
            continue
        xlsx = matches[0]
        if args.only and args.only.lower() not in xlsx.stem.lower():
            continue
        parsed = {tc["e2e_id"]: tc for tc in parse_workbook(xlsx)}
        print(f"[fix] {xlsx.relative_to(args.input)}: {len(entries)} log entries, {len(parsed)} parsed")
        for r in entries:
            tc = parsed.get(r["e2e_id"])
            if not tc:
                print(f"  - {r['qmetry_key']} {r['e2e_id']}: not in workbook now")
                continue
            iid = r["internal_id"]
            try:
                qtm(f"/testcases/{iid}/versions/1", method="PUT",
                    body={"precondition": str(tc["notes"]) if tc["notes"] else ""})
                qtm(f"/testcases/{iid}/versions/1/teststeps", method="DELETE",
                    body={"deleteAll": True})
                if tc["steps"]:
                    qtm(f"/testcases/{iid}/versions/1/teststeps", method="POST", body=tc["steps"])
                fixed += 1
                print(f"  ~ {r['qmetry_key']} {r['e2e_id']}: {len(tc['steps'])} steps")
            except Exception as e:
                failed += 1
                print(f"  ! {r['qmetry_key']} {r['e2e_id']}: {e}")
    print(f"\nfix-existing done. fixed={fixed} failed={failed}")


def run_import(args, log_path: Path):
    if not args.input.exists():
        sys.exit(f"Input directory not found: {args.input}")

    seen = load_log(log_path)
    root_name = args.root_folder_name or args.input.name
    root_id = None if args.dry_run else find_or_create_folder(args.project_id, root_name, args.parent_folder_id)
    if root_id:
        print(f"[folder] {root_name} -> {root_id}")

    category_dirs = sorted(p for p in args.input.iterdir() if p.is_dir())
    created = skipped = failed = 0

    for cat_dir in category_dirs:
        cat_id = None if args.dry_run else find_or_create_folder(args.project_id, cat_dir.name, root_id)
        if cat_id:
            print(f"[folder]   {cat_dir.name} -> {cat_id}")

        for xlsx in sorted(cat_dir.glob("*.xlsx")):
            if xlsx.name.startswith("~$") or xlsx.name.startswith("."):
                continue
            stem = xlsx.stem.replace("_E2E_TestCases", "")
            if args.only and args.only.lower() not in stem.lower():
                continue
            if args.dry_run:
                tcs = list(parse_workbook(xlsx))
                print(f"[dry] {cat_dir.name}/{stem}: {len(tcs)} cases, "
                      f"{sum(len(t['steps']) for t in tcs)} steps")
                for t in tcs:
                    print(f"   - {t['e2e_id']} | pri={t['priority_id']} | steps={len(t['steps'])} | {t['summary'][:80]}")
                continue
            wb_folder_id = find_or_create_folder(args.project_id, stem, cat_id)
            print(f"[folder]     {stem} -> {wb_folder_id}")

            for tc in parse_workbook(xlsx):
                key_pair = (xlsx.name, tc["e2e_id"])
                if key_pair in seen:
                    skipped += 1
                    continue
                try:
                    summary = f"{tc['e2e_id']} | {tc['summary']}"
                    if len(summary) > TC_SUMMARY_MAX:
                        summary = summary[:TC_SUMMARY_MAX - 3] + "..."
                    is_api = str(tc["test_type"]).strip().upper() == "API"
                    body = {
                        "projectId": args.project_id,
                        "folderId": wb_folder_id,
                        "summary": summary,
                        "priority": tc["priority_id"],
                        "status": args.status_id,
                    }
                    if args.component_ids:
                        body["components"] = list(args.component_ids)
                    if args.api_test_field_id:
                        body["customFields"] = [{
                            "id": args.api_test_field_id,
                            "value": args.api_test_yes if is_api else args.api_test_no,
                        }]
                    if tc["notes"]:
                        body["precondition"] = str(tc["notes"])

                    res = qtm("/testcases", method="POST", body=body)
                    key, iid = res["key"], res["id"]
                    if tc["steps"]:
                        qtm(f"/testcases/{iid}/versions/1/teststeps", method="POST", body=tc["steps"])
                    append_log(log_path, xlsx.name, tc["e2e_id"], key, iid, "ok")
                    created += 1
                    print(f"  + {key}  {tc['e2e_id']}  ({len(tc['steps'])} steps)")
                except Exception as e:
                    append_log(log_path, xlsx.name, tc["e2e_id"], "", "", "failed", str(e)[:300])
                    failed += 1
                    print(f"  ! FAILED {tc['e2e_id']}: {e}")

    print(f"\nDone. created={created} skipped={skipped} failed={failed}")
    print(f"Log: {log_path}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description="Import xlsx test cases into QMetry.")
    p.add_argument("--input", type=Path, default=DEFAULT_INPUT_DIR,
                   help=f"Root input directory (default: {DEFAULT_INPUT_DIR})")
    p.add_argument("--parent-folder-id", type=int, default=DEFAULT_PARENT_FOLDER_ID,
                   help="QMetry parent folder ID under which the tree is created")
    p.add_argument("--root-folder-name",
                   help="Name of the top folder to create under --parent-folder-id (default: basename of --input)")
    p.add_argument("--project-id", type=int, default=DEFAULT_PROJECT_ID)
    p.add_argument("--component-id", dest="component_ids", type=int, action="append",
                   default=None,
                   help="Component ID to attach (repeat for multiple). Omit to attach none. "
                        f"Default: {DEFAULT_COMPONENT_IDS}")
    p.add_argument("--status-id", type=int, default=DEFAULT_STATUS_ID,
                   help=f"Status ID for new test cases (default: {DEFAULT_STATUS_ID} = Done)")
    p.add_argument("--api-test-field-id", default=DEFAULT_API_TEST_FIELD_ID,
                   help="QMetry custom field id for 'API Test' radio (empty to skip)")
    p.add_argument("--api-test-yes", default=DEFAULT_API_TEST_OPTION_YES)
    p.add_argument("--api-test-no", default=DEFAULT_API_TEST_OPTION_NO)
    p.add_argument("--log", type=Path,
                   help="Path to log CSV (default: scripts/import-<input-dirname>.log.csv)")
    p.add_argument("--only", help="Substring filter on workbook stem (case-insensitive)")
    p.add_argument("--dry-run", action="store_true",
                   help="Parse and print, do not create anything in QMetry")
    p.add_argument("--fix-existing", action="store_true",
                   help="Re-write precondition + steps for previously imported cases (uses --log)")
    args = p.parse_args()

    if args.component_ids is None:
        args.component_ids = list(DEFAULT_COMPONENT_IDS)
    if not args.log:
        slug = re.sub(r"[^A-Za-z0-9._-]+", "-", args.input.name).strip("-").lower() or "import"
        args.log = Path(__file__).parent / f"import-{slug}.log.csv"

    # Auth — env only (never put the API key on the CLI)
    api_key = os.environ.get("QTM4J_API_KEY")
    if not api_key and not args.dry_run:
        sys.exit("ERROR: QTM4J_API_KEY env var is required (not needed for --dry-run)")
    region = os.environ.get("QTM4J_REGION", "US")
    base_url = "https://syd-qtmcloud.qmetry.com" if region == "AU" else "https://qtmcloud.qmetry.com"
    CFG.api_key = api_key or ""
    CFG.api_base = f"{base_url}/rest/api/latest"

    if args.fix_existing:
        fix_existing(args, args.log)
    else:
        run_import(args, args.log)


if __name__ == "__main__":
    main()
